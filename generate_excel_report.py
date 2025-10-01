#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成Excel综合统计报告
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def format_excel_sheet(ws, title):
    """格式化Excel工作表"""
    # 设置标题行样式
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 冻结首行
    ws.freeze_panes = 'A2'


def create_summary_sheet(writer, workbook):
    """创建汇总说明表"""
    summary_data = {
        '统计维度': [
            '场景+任务+Ring',
            '场景+Ring',
            '任务+Ring',
            '模型+任务',
            '模型+场景',
            '模型+Ring',
            '场景×任务×Ring×模型详细对比',
            '最佳模型推荐',
            '原始数据汇总'
        ],
        '工作表名称': [
            '按场景任务Ring统计',
            '按场景Ring统计',
            '按任务Ring统计',
            '按模型任务统计',
            '按模型场景统计',
            '按模型Ring统计',
            '详细模型对比',
            '最佳模型',
            '原始汇总数据'
        ],
        '说明': [
            '按场景、任务、Ring分组，跨模型统计各指标的均值和标准差',
            '按场景和Ring分组，跨任务和模型统计',
            '按任务和Ring分组，跨场景和模型统计',
            '按模型和任务分组，跨场景和Ring统计',
            '按模型和场景分组，跨任务和Ring统计',
            '按模型和Ring分组，跨场景和任务统计',
            '透视表：展示不同场景、任务、Ring下各模型的性能对比',
            '按场景、任务、Ring组合推荐MAE最优的模型',
            '所有实验的原始汇总数据'
        ]
    }
    
    df = pd.DataFrame(summary_data)
    df.to_excel(writer, sheet_name='说明', index=False)
    
    ws = writer.sheets['说明']
    format_excel_sheet(ws, '统计说明')
    
    # 添加额外信息
    ws['A11'] = '数据来源'
    ws['B11'] = '/root/RingTool/csv/'
    ws['A12'] = '统计日期'
    ws['B12'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A13'] = '模型类型'
    ws['B13'] = 'resnet, inception_time, mamba2, transformer'
    ws['A14'] = 'Ring类型'
    ws['B14'] = 'ring1, ring2'
    ws['A15'] = '任务类型'
    ws['B15'] = 'hr, resp_rr, BP_dia, BP_sys, spo2'
    ws['A16'] = '场景类型'
    ws['B16'] = 'motion, spo2, stationary'
    
    for row in range(11, 17):
        ws[f'A{row}'].font = Font(bold=True)


def main():
    # 创建Excel写入器
    output_file = 'comprehensive_statistics_report.xlsx'
    
    print("=" * 80)
    print("               📊 生成Excel综合统计报告")
    print("=" * 80)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. 创建说明表
        create_summary_sheet(writer, writer.book)
        print("✅ 说明")
        
        # 2. 读取并写入各统计表
        files_to_process = [
            ('stats_by_scenario_task_ring.csv', '按场景任务Ring统计'),
            ('stats_by_scenario_ring.csv', '按场景Ring统计'),
            ('stats_by_task_ring.csv', '按任务Ring统计'),
            ('stats_by_model_task.csv', '按模型任务统计'),
            ('stats_by_model_scenario.csv', '按模型场景统计'),
            ('stats_by_model_ring.csv', '按模型Ring统计'),
            ('best_models_by_scenario_task_ring.csv', '最佳模型'),
            ('all_results_summary.csv', '原始汇总数据'),
            ('detailed_model_comparison.csv', '详细模型对比')
        ]
        
        for csv_file, sheet_name in files_to_process:
            try:
                df = pd.read_csv(csv_file)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                format_excel_sheet(ws, sheet_name)
                print(f"✅ {sheet_name}")
            except Exception as e:
                print(f"⚠️  {sheet_name}: {e}")
    
    print("\n" + "=" * 80)
    print(f"✅ Excel报告已生成: {output_file}")
    print("=" * 80)
    print("\n📋 包含的工作表：")
    print("  1. 说明 - 统计说明和元信息")
    print("  2. 按场景任务Ring统计")
    print("  3. 按场景Ring统计")
    print("  4. 按任务Ring统计")
    print("  5. 按模型任务统计")
    print("  6. 按模型场景统计")
    print("  7. 按模型Ring统计")
    print("  8. 最佳模型")
    print("  9. 原始汇总数据")
    print(" 10. 详细模型对比")
    print("=" * 80)


if __name__ == '__main__':
    main()


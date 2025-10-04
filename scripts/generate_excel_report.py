#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Report Generator for Statistical Analysis
Consolidates all statistical tables into a single workbook
"""

import pandas as pd
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def format_sheet(ws):
    """Apply professional formatting to worksheet"""
    if not HAS_OPENPYXL:
        return
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'


def create_overview_sheet(writer):
    """Create README worksheet with table descriptions"""
    overview = pd.DataFrame({
        'Table': [
            '== PRIMARY TABLES ==',
            'Table 1: By Channel',
            '  All-IR',
            '  All-IRIMU', 
            '  All-IRRED',
            '  All-IRREDIMU',
            '',
            'Table 2: By Scenario Group',
            '  Motion-[CHANNEL]',
            '  Spo2-[CHANNEL]',
            '  Stationary-[CHANNEL]',
            '',
            'Table 3: Channel Comparison',
            '',
            'Table 4: Best Models',
            '',
            '== RAW DATA ==',
            'Raw Aggregated',
            'Raw Fold Level',
        ],
        'Description': [
            '',
            'Performance across all scenarios for each channel',
            'IR channel only',
            'IRIMU channel',
            'IRRED channel', 
            'IRREDIMU channel',
            '',
            'Performance by scenario group (12 sheets total)',
            'Motion scenarios: deepsquat, striding',
            'SPO2 scenario',
            'Stationary: sitting, talking, shaking_head, standing',
            '',
            'Side-by-side comparison of all channels',
            '',
            'Recommended model for each configuration',
            '',
            '',
            'Aggregated 5-fold data (basis for all tables)',
            'Individual fold results',
        ]
    })
    
    overview.to_excel(writer, sheet_name='README', index=False)
    
    if HAS_OPENPYXL:
        ws = writer.sheets['README']
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60


def main():
    base_dir = Path(__file__).parent.parent
    stats_dir = base_dir / 'statistics_results'
    output_file = stats_dir / 'statistical_report.xlsx'
    
    print("="*80)
    print("Generating Excel Report")
    print("="*80)
    
    if not stats_dir.exists():
        print(f"\nError: Directory not found - {stats_dir}")
        print("Run calculate_metrics_statistics.py first")
        return
    
    if not HAS_OPENPYXL:
        print("\nError: openpyxl not installed")
        print("Install: pip install openpyxl")
        return
    
    print(f"\nInput: {stats_dir}")
    print(f"Output: {output_file}\n")
    
    tables = []
    
    # Table 1: By channel
    for channel in ['ir', 'irimu', 'irred', 'irredimu']:
        tables.append((f'table1_{channel}_weighted.csv', f'All-{channel.upper()}'))
    
    # Table 2: By scenario group x channel
    for scenario in ['motion', 'spo2', 'stationary']:
        for channel in ['ir', 'irimu', 'irred', 'irredimu']:
            tables.append((f'table2_{scenario}_{channel}.csv', f'{scenario.title()}-{channel.upper()}'))
    
    # Other tables
    tables.extend([
        ('table3_channel_comparison.csv', 'Channel Comparison'),
        ('table4_best_models.csv', 'Best Models'),
        ('raw_aggregated.csv', 'Raw Aggregated'),
        ('raw_fold_level.csv', 'Raw Fold Level'),
    ])
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            create_overview_sheet(writer)
            print("  README")
            
            for csv_file, sheet_name in tables:
                csv_path = stats_dir / csv_file
                
                if not csv_path.exists():
                    continue
                
                try:
                    df = pd.read_csv(csv_path)
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                    ws = writer.sheets[sheet_name[:31]]
                    format_sheet(ws)
                    print(f"  {sheet_name}")
                except Exception as e:
                    print(f"  Warning: {sheet_name} - {e}")
        
        print("\n" + "="*80)
        print(f"Excel report generated: {output_file}")
        print("="*80)
        
    except Exception as e:
        print(f"\nError: {e}")


if __name__ == '__main__':
    main()
